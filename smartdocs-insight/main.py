"""
SmartDocs Insight – main GUI module.

This module implements the desktop workflow for:
- Selecting invoice CSV/PDF files belonging to the same invoice,
- Parsing COO and weight information from CSV line items,
- Detecting missing or malformed COO entries,
- Aggregating PO/Box 5 and invoice references,
- Generating an email body for Certificate of Origin (CoO) requests,
- Exporting an Excel-based tracking sheet.

Author: Damian Komorowski
"""

from __future__ import annotations

import csv
import json
import logging
import os
import re
import time
import webbrowser
from datetime import datetime
from typing import List, Tuple

import fitz  # PyMuPDF
import openpyxl
import pycountry
import pyperclip
import tkinter as tk
from tkinter import filedialog, messagebox, ttk


# ---------------------------------------------------------------------------
# Configuration constants – safe defaults for public GitHub
# ---------------------------------------------------------------------------

# TODO: If you use this in a corporate context, configure these via environment
#       variables or a config file instead of hardcoding.

DEFAULT_COO_REQUEST_EMAIL = os.getenv(
    "SMARTDOCS_COO_EMAIL", "classification@example.com"
)
DEFAULT_SUPPORT_EMAIL = os.getenv(
    "SMARTDOCS_SUPPORT_EMAIL", "support@example.com"
)
DEFAULT_SUPPORT_PHONE_URL = os.getenv(
    "SMARTDOCS_SUPPORT_PHONE_URL", "https://example.com/contact"
)
DEFAULT_DOC_URL = os.getenv(
    "SMARTDOCS_DOC_URL", "https://example.com/internal-docs"
)

# Path to preferences file in the same folder as this module
prefs_file_path = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "user_preferences.json"
)


def _ensure_prefs_file_exists() -> dict:
    """
    Load preferences from JSON file. If the file does not exist or is invalid,
    create a default one with dark_mode set to True.

    Returns:
        dict: Preference dictionary.
    """
    try:
        if os.path.exists(prefs_file_path):
            with open(prefs_file_path, "r", encoding="utf-8") as f:
                content = f.read().strip()
                prefs = json.loads(content) if content else {}
        else:
            prefs = {}
    except (FileNotFoundError, json.JSONDecodeError):
        prefs = {}

    if "dark_mode" not in prefs:
        prefs["dark_mode"] = True

    # Persist back (create or fix invalid file)
    try:
        with open(prefs_file_path, "w", encoding="utf-8") as f:
            json.dump(prefs, f, indent=4)
    except Exception as e:
        print(f"⚠️ Failed to initialize preferences file: {e}")

    return prefs


_prefs_cache = _ensure_prefs_file_exists()
initial_dark_mode: bool = bool(_prefs_cache.get("dark_mode", True))

# EU ISO country codes (alpha-2)
EU_CODES = {
    "AT",
    "BE",
    "BG",
    "HR",
    "CY",
    "CZ",
    "DK",
    "EE",
    "FI",
    "FR",
    "DE",
    "GR",
    "HU",
    "IE",
    "IT",
    "LV",
    "LT",
    "LU",
    "MT",
    "NL",
    "PL",
    "PT",
    "RO",
    "SK",
    "SI",
    "ES",
    "SE",
}


# ---------------------------------------------------------------------------
# Helper functions – country & invoice analysis
# ---------------------------------------------------------------------------


def get_country_name(code: str) -> str:
    """
    Convert a two-letter country code into a human-readable country name.

    Args:
        code: ISO alpha-2 country code (e.g. 'PL', 'MX').

    Returns:
        str: Country name or the original code if lookup fails.
    """
    try:
        upper = code.upper()
        if upper == "KR":
            return "Republic of Korea"
        country = pycountry.countries.get(alpha_2=upper)
        return country.name if country else code
    except Exception:
        return code


def analyze_invoice(
    file_path: str,
) -> Tuple[List[str], List[str], float, str, str, List[str]]:
    """
    Parse an invoice CSV file and return structured classification data.

    Returns:
        tuple:
            non_eu_items: list[str]
            eu_items: list[str]
            total_non_eu_weight_kg: float
            box5_reference: str
            invoice_reference: str
            missing_coo_alerts: list[str]

    Notes:
        - Reads B1 as the primary invoice reference (BOM-safe).
        - Accepts HEADER/INVOICE/INV in A1 as a fallback (BOM stripped).
        - Handles COO/weight formats with units: KG / KGS / G / GRAM(S).
        - Processes rows with 'ITEM' in column A (case-insensitive).
    """

    def _norm(s: str) -> str:
        # Trim, remove UTF-8 BOM if present, uppercase for comparisons
        return (s or "").strip().lstrip("\ufeff").upper()

    non_eu_items: List[str] = []
    eu_items: List[str] = []
    total_non_eu_weight = 0.0
    invoice_reference = "Unknown"
    box5_values = set()
    missing_coo_alerts: List[str] = []

    # Regex for COO/weight information, e.g.: "/ MX / 2.497 KG"
    coo_weight_pattern = re.compile(
        r"/\s*(?P<coo>[A-Za-z]{2})\s*/\s*(?P<weight>[\d.,]+)\s*(?P<unit>KG|KGS|G|GRAMS?)\b",
        re.IGNORECASE,
    )

    try:
        # Use utf-8-sig so BOM at file start is stripped automatically
        with open(file_path, newline="", encoding="utf-8-sig") as csvfile:
            reader = csv.reader(csvfile)
            lines = list(reader)

        # ---- Invoice reference (prefer B1; guarded header check) ----
        if len(lines) > 0 and len(lines[0]) > 1:
            invoice_reference = (lines[0][1] or "").strip()

            if not invoice_reference:
                first_cell = _norm(lines[0][0])
                if first_cell in {"HEADER", "INVOICE", "INV"}:
                    invoice_reference = (lines[0][1] or "").strip()

        # ---- Iterate all rows; only handle those starting with 'ITEM' ----
        for row in lines:
            if not row:
                continue

            if _norm(row[0]) != "ITEM":
                continue

            # Defensive extraction with column guards
            catalog = row[11].strip() if len(row) > 11 and row[11] else "Unknown"
            description = row[5].strip() if len(row) > 5 and row[5] else "Unknown"

            # Box 5 is column J (index 9)
            if len(row) > 9 and row[9]:
                raw_box5 = row[9].strip()
                if raw_box5:
                    box5_values.add(raw_box5)

            # COO/weight field is column Q (index 16) by the sample CSV
            coo_weight_field = row[16].strip() if len(row) > 16 and row[16] else ""
            match = coo_weight_pattern.search(coo_weight_field)

            # If slashes are present but COO isn't matched → alert as "Missing COO"
            if not match:
                if "/" in coo_weight_field:
                    line_number = row[1].strip() if len(row) > 1 and row[1] else "Unknown"
                    weight_info = coo_weight_field
                    if "/" in weight_info:
                        # Take the last part after the last slash (likely the "weight unit" chunk)
                        weight_info = weight_info.split("/")[-1].strip()
                    weight_info = " ".join(weight_info.split())  # normalize spaces

                    missing_coo_alerts.append(
                        f"⚠️ Missing COO → Line {line_number}\n"
                        f" Product: {catalog}\n"
                        f" Desc: {description}\n"
                        f" Weight: {weight_info}"
                    )
                # Skip this line for COO processing if not matched
                continue

            # Extract COO + weight
            coo = match.group("coo").strip().upper()
            weight_str = match.group("weight").replace(",", "").strip()
            unit = match.group("unit").strip().upper()

            # Convert weight to KG
            try:
                weight_value = float(weight_str)
                if unit.startswith("G"):  # "G" or "GRAM(S)"
                    weight_value /= 1000.0
            except ValueError:
                weight_value = 0.0

            # Determine country name
            special_names = {"KR": "Republic of Korea"}
            country_name = {
                "MX": "Mexico",
                "MY": "Malaysia",
                "PL": "Poland",
            }.get(coo, special_names.get(coo, get_country_name(coo)))

            item_text = f"{catalog}, {description}, {country_name}"

            if coo in EU_CODES:
                eu_items.append(item_text)
            else:
                non_eu_items.append(item_text)
                total_non_eu_weight += weight_value

        box5_reference = ", ".join(sorted(box5_values)) if box5_values else "Unknown"

        return (
            non_eu_items,
            eu_items,
            round(total_non_eu_weight, 3),
            box5_reference,
            invoice_reference or "Unknown",
            missing_coo_alerts,
        )

    except (PermissionError, OSError, UnicodeDecodeError, csv.Error) as e:
        return ([], [], 0.0, "Unknown", "Unknown", [f"Error during analysis: {e}"])
    except Exception as e:
        return ([], [], 0.0, "Unknown", "Unknown", [f"Error during analysis: {e}"])


def extract_csv_metadata(
    file_path: str,
) -> Tuple[str, str, List[Tuple[str, str]], str | None]:
    """
    Extract metadata from an invoice CSV file.

    Returns:
        customer_name: str
        invoice_number: str
        po_so_pairs: list[(PO, SO)]
        error: str | None
    """
    customer_name = ""
    invoice_number = ""
    po_so_pairs: List[Tuple[str, str]] = []
    error: str | None = None

    try:
        with open(file_path, newline="", encoding="utf-8") as csvfile:
            reader = csv.reader(csvfile)
            lines = list(reader)

        # Extract invoice number and customer name from header lines
        if len(lines) > 0 and len(lines[0]) > 1:
            invoice_number = (lines[0][1] or "").strip()
        if len(lines) > 1 and len(lines[1]) > 1:
            customer_name = (lines[1][1] or "").strip()

        # Extract PO and SO pairs from columns J and N (index 9 and 13)
        seen = set()
        for row in lines[7:]:
            if not row:
                continue

            po = row[9].strip() if len(row) > 9 and row[9] else ""
            so_full = row[13].strip() if len(row) > 13 and row[13] else ""

            if not po or not so_full:
                continue

            # Remove suffix like "-000010"
            so = so_full.split("-")[0].strip()
            key = (po, so)

            if key not in seen:
                seen.add(key)
                po_so_pairs.append(key)

    except (PermissionError, OSError, UnicodeDecodeError, csv.Error) as e:
        error = f"Error reading CSV: {e}"
    except Exception as e:
        error = f"Error reading CSV: {e}"

    return customer_name, invoice_number, po_so_pairs, error


def extract_bpid_from_pdf(file_path: str) -> str:
    """
    Extract BPID / customer ID from an invoice PDF.

    The function scans for known language variants of "customer ID" labels and
    returns the first numeric (or non-empty) line directly following the label.

    Returns:
        str: Extracted BPID as string, or an empty string if not found.
    """
    customer_id_labels = [
        "Vs. Codice Cliente",
        "Your Customer ID",
        "N° Compte Client",
        "Kundennummer",
        "Uw Klantnummer",
        "Nº Cliente",
    ]

    try:
        doc = fitz.open(file_path)
    except Exception:
        return ""

    try:
        for page in doc:
            try:
                lines = page.get_text().splitlines()
            except Exception:
                continue

            for i, line in enumerate(lines):
                for label in customer_id_labels:
                    if label in line:
                        # Search for the next non-empty line; prefer numeric-only if available
                        for j in range(i + 1, len(lines)):
                            next_line = (lines[j] or "").strip()
                            if not next_line:
                                continue
                            if next_line.isdigit():
                                return next_line
                            # If not purely digits but present, return it anyway
                            return next_line

    except Exception:
        return ""
    finally:
        try:
            doc.close()
        except Exception:
            pass

    return ""


# ---------------------------------------------------------------------------
# Invoice counter & preferences helpers
# ---------------------------------------------------------------------------


def increment_invoice_analysis_counter() -> None:
    """
    Increment the invoice analysis counter stored in the preferences file.
    """
    try:
        if os.path.exists(prefs_file_path):
            with open(prefs_file_path, "r", encoding="utf-8") as f:
                prefs = json.load(f)
        else:
            prefs = {}

        prefs["invoice_analysis_count"] = prefs.get("invoice_analysis_count", 0) + 1

        with open(prefs_file_path, "w", encoding="utf-8") as f:
            json.dump(prefs, f, indent=4)
    except Exception as e:
        print(f"⚠️ Failed to increment invoice analysis counter: {e}")


def get_invoice_analysis_count() -> int:
    """
    Read the invoice analysis counter from the preferences file.

    Returns:
        int: Number of invoices processed so far.
    """
    try:
        if os.path.exists(prefs_file_path):
            with open(prefs_file_path, "r", encoding="utf-8") as f:
                prefs = json.load(f)
            return int(prefs.get("invoice_analysis_count", 0))
        return 0
    except Exception as e:
        print(f"⚠️ Failed to read invoice analysis count: {e}")
        return 0


def reset_invoice_analysis_counter() -> bool:
    """
    Reset the invoice analysis counter to zero.

    Returns:
        bool: True if successful, otherwise False.
    """
    try:
        if os.path.exists(prefs_file_path):
            with open(prefs_file_path, "r", encoding="utf-8") as f:
                prefs = json.load(f)
        else:
            prefs = {}

        prefs["invoice_analysis_count"] = 0

        with open(prefs_file_path, "w", encoding="utf-8") as f:
            json.dump(prefs, f, indent=4)

        return True
    except Exception as e:
        print(f"⚠️ Failed to reset counter: {e}")
        return False


def save_dark_mode_preference(dark_mode_value: bool) -> None:
    """
    Persist the dark_mode preference while keeping all other keys intact.
    """
    try:
        if os.path.exists(prefs_file_path):
            with open(prefs_file_path, "r", encoding="utf-8") as f:
                prefs = json.load(f)
        else:
            prefs = {}

        prefs["dark_mode"] = bool(dark_mode_value)

        with open(prefs_file_path, "w", encoding="utf-8") as f:
            json.dump(prefs, f, indent=4)
    except Exception as e:
        print(f"⚠️ Failed to save dark mode preference: {e}")


# ---------------------------------------------------------------------------
# GUI helpers – responsible person dialog & Excel export
# ---------------------------------------------------------------------------


def get_responsible_person() -> Tuple[str, bool]:
    """
    Show a modal dialog to capture the responsible person and whether
    the generated Excel file should be opened automatically.

    Returns:
        tuple:
            responsible_person: str
            open_excel: bool
    """
    # Honor global dark-mode preference if available
    try:
        dark_mode_enabled = bool(initial_dark_mode)
    except NameError:
        dark_mode_enabled = True

    bg = "#1e1e1e" if dark_mode_enabled else "#f4f4f4"
    fg = "#ffffff" if dark_mode_enabled else "#000000"
    accent = "#3399ff" if dark_mode_enabled else "#0078D7"
    subtle = "#cccccc" if dark_mode_enabled else "#666666"
    btn_bg = "#444444" if dark_mode_enabled else "#dddddd"

    window = tk.Toplevel()
    window.title("Tracking List Generation")
    window.configure(padx=24, pady=24, bg=bg)
    window.resizable(True, True)
    window.transient()
    window.grab_set()
    window.update_idletasks()

    selected_name = tk.StringVar()
    open_excel_choice = tk.StringVar(value="YES")

    def on_cancel() -> None:
        selected_name.set("")
        open_excel_choice.set("NO")
        window.destroy()

    window.protocol("WM_DELETE_WINDOW", on_cancel)
    window.bind("<Escape>", lambda e: on_cancel())

    header = tk.Label(
        window,
        text="👤 Select Responsible Person",
        font=("Segoe UI", 14, "bold"),
        fg=fg,
        bg=bg,
    )
    header.grid(row=0, column=0, columnspan=2, pady=(0, 12), sticky="w")

    def hover_effect(widget: tk.Widget, enter_color: str, leave_color: str) -> None:
        widget.bind("<Enter>", lambda e: widget.config(bg=enter_color))
        widget.bind("<Leave>", lambda e: widget.config(bg=leave_color))

    # TODO: For public GitHub, keep these generic. Customize in your local fork.
    names = ["Team Member 1", "Team Member 2", "Team Member 3"]
    for i, name in enumerate(names, start=1):
        btn = tk.Button(
            window,
            text=name,
            width=28,
            font=("Segoe UI", 10, "bold"),
            bg=btn_bg,
            fg=fg,
            relief="flat",
            command=lambda n=name: (selected_name.set(n), window.destroy()),
        )
        btn.grid(row=i, column=0, pady=3, sticky="w")
        hover_effect(btn, accent, btn_bg)

    manual_label = tk.Label(
        window,
        text="✍️ Or enter your name:",
        font=("Segoe UI", 10),
        fg=subtle,
        bg=bg,
    )
    manual_label.grid(row=5, column=0, pady=(14, 4), sticky="w")

    entry = tk.Entry(window, width=32, font=("Segoe UI", 10))
    entry.grid(row=6, column=0, pady=4, sticky="w")
    entry.focus_set()

    feedback = tk.Label(
        window, text="", font=("Segoe UI", 9), fg="#ff6666", bg=bg
    )
    feedback.grid(row=7, column=0, sticky="w")

    def do_confirm() -> None:
        name_val = (selected_name.get() or entry.get()).strip()
        if not name_val:
            feedback.config(text="Please select a name or type your name.")
            entry.focus_set()
            return
        selected_name.set(name_val)
        window.destroy()

    confirm_btn = tk.Button(
        window,
        text="Confirm",
        font=("Segoe UI", 10, "bold"),
        bg=accent,
        fg="white",
        relief="flat",
        command=do_confirm,
    )
    confirm_btn.grid(row=8, column=0, pady=(10, 0), sticky="w")
    hover_effect(
        confirm_btn,
        "#5ab0ff" if dark_mode_enabled else "#3399ff",
        accent,
    )

    excel_label = tk.Label(
        window,
        text="📊 Open auto-generated\nExcel file?",
        font=("Segoe UI", 10, "bold"),
        fg=fg,
        bg=bg,
    )
    excel_label.grid(row=1, column=1, rowspan=1, padx=18, sticky="nw")

    yes_radio = tk.Radiobutton(
        window,
        text="YES",
        variable=open_excel_choice,
        value="YES",
        font=("Segoe UI", 10),
        bg=bg,
        fg=fg,
        selectcolor=bg,
    )
    yes_radio.grid(row=2, column=1, sticky="nw", padx=18)

    no_radio = tk.Radiobutton(
        window,
        text="NO",
        variable=open_excel_choice,
        value="NO",
        font=("Segoe UI", 10),
        bg=bg,
        fg=fg,
        selectcolor=bg,
    )
    no_radio.grid(row=3, column=1, sticky="nw", padx=18)

    window.bind("<Return>", lambda e: do_confirm())

    window.update_idletasks()
    w, h = 520, 320
    try:
        w = max(w, window.winfo_width() + 40)
        h = max(h, window.winfo_height() + 20)
    except Exception:
        pass

    window.geometry(f"{w}x{h}+0+0")
    window.wait_window()

    return selected_name.get(), (open_excel_choice.get() == "YES")


def run_metadata_export(
    file_paths: List[str],
    responsible_person: str,
    open_excel: bool,
) -> None:
    """
    Build an Excel-based tracking sheet using selected CSV and PDF files.

    Args:
        file_paths: List of selected CSV and PDF file paths.
        responsible_person: Name of the responsible person.
        open_excel: Whether to open Excel automatically after export.
    """
    try:
        logger = logging.getLogger("SmartDocs.Export")
        if not logger.handlers:
            log_path = os.path.join(
                os.path.dirname(os.path.abspath(__file__)), "SmartDocs.log"
            )
            logging.basicConfig(
                filename=log_path,
                level=logging.INFO,
                format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
            )
        debug_mode = os.getenv("SMARTDOCS_DEBUG", "0") == "1"
        if debug_mode:
            logger.setLevel(logging.DEBUG)
            logger.debug("run_metadata_export() started.")
    except Exception:
        logger = None

    csv_files = [f for f in file_paths if f.lower().endswith(".csv")]
    pdf_files = [f for f in file_paths if f.lower().endswith(".pdf")]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TemporaryData"

    headers = [
        "Lp.",
        "CUSTOMER NAME",
        "Customer number",
        "REQUEST TYPE",
        "DOC",
        "Invoice",
        "PO",
        "SO",
        "Status",
        "LT",
        "DATE RECEIVED",
        "DATE REQUESTED",
        "DATE COMPLETED",
        "Incident Number",
        "Comments",
        "RESPONSIBLE PERSON",
    ]
    ws.append(headers)

    status_text = "Requested"
    lt_text = "Invoice Created"
    today = datetime.today().strftime("%Y-%m-%d")
    counter = 1

    pdf_bpid = extract_bpid_from_pdf(pdf_files[0]) if pdf_files else ""

    if logger:
        try:
            logger.info(
                "Export initiated. CSV files: %d, PDF files: %d, BPID: %s",
                len(csv_files),
                len(pdf_files),
                pdf_bpid,
            )
        except Exception:
            pass

    total_rows_added = 0
    for file_path in csv_files:
        client, inv, po_so_pairs, error = extract_csv_metadata(file_path)
        if logger:
            try:
                logger.debug(
                    "CSV read: %s; invoice=%s; client=%s; pairs=%d; error=%s",
                    os.path.basename(file_path),
                    inv,
                    client,
                    len(po_so_pairs),
                    error,
                )
            except Exception:
                pass

        seen_pairs = set()
        unique_po_so_pairs: List[Tuple[str, str]] = []
        for po, so in po_so_pairs:
            if (po, so) not in seen_pairs:
                seen_pairs.add((po, so))
                unique_po_so_pairs.append((po, so))

        for po, so in unique_po_so_pairs:
            row_data = [
                counter,
                client,
                pdf_bpid,
                "STANDARD",
                "DOC",
                inv,
                po,
                so,
                status_text,
                lt_text,
                today,
                today,
                "",
                "",
                "",
                responsible_person,
            ]
            ws.append(row_data)
            counter += 1
            total_rows_added += 1

    try:
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2
    except Exception as e:
        if logger:
            logger.warning("Column autosize warning: %s", e)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    temp_filename = f"TEMP_Metadata_Export_{timestamp}.xlsx"

    if logger:
        try:
            logger.info(
                "Export target path: %s; rows: %d", temp_filename, total_rows_added
            )
        except Exception:
            pass

    try:
        wb.save(temp_filename)
    except PermissionError as e:
        messagebox.showerror(
            "Cannot Save File",
            "The export file could not be saved (file may be in use or locked).\n\n"
            f"Path: {temp_filename}\n\nDetails:\n{e}",
        )
        if logger:
            logger.error("Save failed due to PermissionError: %s", e)
        return
    except Exception as e:
        messagebox.showerror(
            "Save Error",
            "An unexpected error occurred while saving the export file.\n\n"
            f"Path: {temp_filename}\n\nDetails:\n{e}",
        )
        if logger:
            logger.error("Save failed due to unexpected error: %s", e)
        return

    clipboard_data = ""
    for row in ws.iter_rows(
        min_row=2,
        max_row=ws.max_row,
        min_col=2,
        max_col=ws.max_column,
        values_only=True,
    ):
        if row[0] not in (None, ""):
            clipboard_data += "\t".join(
                [str(cell) if cell is not None else "" for cell in row]
            ) + "\n"

    try:
        pyperclip.copy(clipboard_data)
        if logger:
            logger.debug("Clipboard copy completed.")
    except Exception as e:
        if logger:
            logger.warning("Clipboard copy warning: %s", e)

    skip_open = os.getenv("SMARTDOCS_SKIP_OPEN", "0") == "1"
    if open_excel and not skip_open:
        try:
            os.startfile(temp_filename)  # type: ignore[attr-defined]
            if logger:
                logger.info("Excel opened: %s", temp_filename)
        except FileNotFoundError as e:
            messagebox.showerror(
                "Open Error",
                "Cannot find the exported file:\n"
                f"{temp_filename}\n\nDetails:\n{e}",
            )
            if logger:
                logger.error("Open failed (FileNotFoundError): %s", e)
        except OSError as e:
            messagebox.showerror(
                "Open Error",
                "Cannot open Excel for:\n"
                f"{temp_filename}\n\nDetails:\n{e}",
            )
            if logger:
                logger.error("Open failed (OSError): %s", e)
        except Exception as e:
            messagebox.showerror(
                "Open Error",
                f"Unexpected error while opening Excel:\n{e}",
            )
            if logger:
                logger.error("Open failed (unexpected): %s", e)
    elif logger:
        logger.info(
            "Open skipped (open_excel=%s, SMARTDOCS_SKIP_OPEN=%s)",
            open_excel,
            skip_open,
        )


# ---------------------------------------------------------------------------
# Main GUI – SmartDocs Insight
# ---------------------------------------------------------------------------


def show_gui() -> None:
    """
    Initialize and run the main SmartDocs Insight window.

    Responsibilities:
        - Configure Tkinter root window (size, theme, DPI awareness).
        - Wire up file selection, dark/light mode and export options.
        - Display email output and processing statistics.
        - Provide access to the Excel-based tracking sheet export.
    """
    root = tk.Tk()
    root.title("SmartDocs Insight")

    root.withdraw()

    try:
        current_scaling = root.tk.call("tk", "scaling")
        if not current_scaling or float(current_scaling) <= 0:
            root.tk.call("tk", "scaling", 1.0)
    except Exception:
        pass

    root.update_idletasks()
    screen_w = root.winfo_screenwidth()
    screen_h = root.winfo_screenheight()
    gui_w = max(960, int(screen_w * 0.40))
    gui_h = max(720, int(screen_h * 0.90))
    root.geometry(f"{gui_w}x{gui_h}+0+0")

    root.minsize(900, 680)
    root.resizable(True, True)
    root.deiconify()

    dark_mode = initial_dark_mode
    enable_excel_export = tk.BooleanVar(value=True)
    selected_color = tk.StringVar(value="#0078D7")

    try:
        if os.path.exists(prefs_file_path):
            with open(prefs_file_path, "r", encoding="utf-8") as f:
                prefs = json.load(f)
            selected_color.set(prefs.get("button_color", "#0078D7"))
        else:
            selected_color.set("#0078D7")
    except Exception:
        selected_color.set("#0078D7")

    style = ttk.Style(root)

    def apply_theme() -> None:
        bg = "#1e1e1e" if dark_mode else "#f4f4f4"
        fg = "#ffffff" if dark_mode else "#333333"
        accent = selected_color.get()
        secondary = "#cccccc" if dark_mode else "#888888"
        text_bg = "#2e2e2e" if dark_mode else "white"
        text_fg = "#ffffff" if dark_mode else "#000000"
        insert_color = "#ffffff" if dark_mode else "#000000"
        button_bg = "#444444" if dark_mode else "#dddddd"
        scrollbar_bg = "#444444" if dark_mode else "#cccccc"
        scrollbar_trough = "#2e2e2e" if dark_mode else "#eeeeee"

        root.configure(bg=bg)

        for frame in [header_frame, button_frame, info_frame, footer_frame]:
            frame.configure(bg=bg)

        label.config(bg=bg, fg=fg)
        subheader.config(bg=bg, fg=secondary)

        button.config(bg=accent, fg="white", activebackground=accent)
        excel_checkbox.config(bg=bg, fg=fg, activebackground=accent, selectcolor=bg)
        doc_button.config(bg=button_bg, fg=fg, activebackground=accent)

        summary_label.config(bg=bg, fg=fg)
        email_label.config(bg=bg, fg=secondary)
        copy_button.config(bg=button_bg, fg=fg, activebackground=accent)

        text_area.config(
            bg=text_bg,
            fg=text_fg,
            insertbackground=insert_color,
        )

        copy_output_button.config(
            bg=button_bg,
            fg=fg,
            activebackground=accent,
        )

        toggle_button.config(
            bg=button_bg,
            fg=fg,
            activebackground=accent,
        )

        processing_label.config(bg=bg, fg=secondary)

        style.theme_use("default")
        style.configure(
            "Vertical.TScrollbar",
            background=scrollbar_bg,
            troughcolor=scrollbar_trough,
            arrowcolor=accent,
            bordercolor=scrollbar_bg,
        )

    def toggle_dark_mode() -> None:
        nonlocal dark_mode
        dark_mode = not dark_mode
        toggle_button.config(text="☀️ Light Mode" if dark_mode else "🌙 Dark Mode")
        apply_theme()
        save_dark_mode_preference(dark_mode)

    def handle_reset_shortcut(event: tk.Event) -> None:
        if reset_invoice_analysis_counter():
            processing_label.config(
                text="Counter reset to 0\nTotal documentation analyzed by the user: 0"
            )

    root.bind("<Shift-R>", handle_reset_shortcut)

    def copy_email() -> None:
        root.clipboard_clear()
        root.clipboard_append(DEFAULT_COO_REQUEST_EMAIL)

    def open_documentation() -> None:
        # TODO: In a corporate setup, point this to your internal documentation URL.
        webbrowser.open(DEFAULT_DOC_URL)

    def select_files() -> None:
        try:
            logger = logging.getLogger("SmartDocs.Select")
            if not logger.handlers:
                log_path = os.path.join(
                    os.path.dirname(os.path.abspath(__file__)), "SmartDocs.log"
                )
                logging.basicConfig(
                    filename=log_path,
                    level=logging.INFO,
                    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
                )
            debug_mode = os.getenv("SMARTDOCS_DEBUG", "0") == "1"
            if debug_mode:
                logger.setLevel(logging.DEBUG)
                logger.debug("select_files() started.")
        except Exception:
            logger = None

        start_time = time.time()

        file_paths = filedialog.askopenfilenames(
            title="Select Invoice CSV & PDF(s)",
            filetypes=[("CSV and PDF Files", "*.csv *.pdf")],
        )

        if not file_paths:
            if logger:
                logger.info("No files selected.")
            return

        csv_files = [f for f in file_paths if f.lower().endswith(".csv")]
        pdf_files = [f for f in file_paths if f.lower().endswith(".pdf")]

        if logger:
            try:
                logger.info("Selected: CSV=%d, PDF=%d", len(csv_files), len(pdf_files))
            except Exception:
                pass

        if csv_files and pdf_files:

            def _sig(path: str) -> str:
                base = os.path.basename(path)
                base = re.sub(r"\(.*?\)", "", base)
                base = os.path.splitext(base)[0]
                base = base.replace("INV_CSV_", "").replace("INV_PDF_D_", "")
                return base.strip()

            csv_sigs = {_sig(f) for f in csv_files}
            pdf_sigs = {_sig(f) for f in pdf_files}

            if csv_sigs != pdf_sigs:
                if logger:
                    logger.warning(
                        "Invoice signature mismatch. CSV=%s, PDF=%s",
                        csv_sigs,
                        pdf_sigs,
                    )
                messagebox.showerror(
                    "Invoice Mismatch",
                    "The selected CSV and PDF files do NOT belong to the same invoice.\n\n"
                    f"CSV invoice signatures detected:\n {', '.join(sorted(csv_sigs))}\n\n"
                    f"PDF invoice signatures detected:\n {', '.join(sorted(pdf_sigs))}\n\n"
                    "Please select files belonging to the SAME invoice number.",
                )
                return

            pdf_bpids = [extract_bpid_from_pdf(f) for f in pdf_files]
            pdf_bpids = [b for b in pdf_bpids if b]

            if len(set(pdf_bpids)) > 1:
                if logger:
                    logger.warning("BPID mismatch detected: %s", pdf_bpids)
                messagebox.showwarning(
                    "BPID Mismatch",
                    "The BPID values extracted from the selected PDF files are not the same.\n"
                    "Please verify that you selected the correct files belonging to the same customer.\n\n",
                )
                return

        if not csv_files or not pdf_files:
            if logger:
                logger.info("Missing file types: either CSV or PDF not provided.")
            messagebox.showwarning(
                "Missing File Types",
                "Please select both CSV and PDF files.",
            )
            return

        all_non_eu_items: List[str] = []
        all_eu_items: List[str] = []
        all_po_numbers: List[str] = []
        all_invoice_numbers: List[str] = []
        total_weight = 0.0

        all_missing_coo: List[str] = []
        file_issues: List[str] = []
        multiple_files = len(csv_files) > 1

        def _show_list_modal(
            title: str,
            lines: List[str],
            bullet: str = "🔸",
        ) -> None:
            if os.getenv("SMARTDOCS_SKIP_ALERTS", "0") == "1":
                if logger:
                    logger.info("Alert skipped (%s); items=%d", title, len(lines))
                return

            import tkinter.font as tkfont

            dark_bg = "#1e1e1e" if initial_dark_mode else "#f4f4f4"
            dark_txt = "#ffffff" if initial_dark_mode else "#000000"
            accent_color = "#3399ff" if initial_dark_mode else "#0078D7"

            safe_lines = [str(l) for l in (lines or ["(no items)"])]

            alert = tk.Toplevel(root)
            alert.title(title)
            alert.configure(bg=dark_bg)
            alert.resizable(True, True)
            alert.transient(root)
            alert.grab_set()

            alert.grid_rowconfigure(1, weight=1)
            alert.grid_columnconfigure(0, weight=1)

            header_label = tk.Label(
                alert,
                text=title,
                font=("Segoe UI", 12, "bold"),
                fg=accent_color,
                bg=dark_bg,
                anchor="w",
                justify="left",
                wraplength=2000,
            )
            header_label.grid(row=0, column=0, sticky="ew", padx=22, pady=(18, 10))

            text_frame = tk.Frame(alert, bg=dark_bg)
            text_frame.grid(row=1, column=0, sticky="nsew", padx=22)

            txt = tk.Text(
                text_frame,
                wrap="word",
                font=("Segoe UI", 10),
                bg="#2e2e2e" if initial_dark_mode else "white",
                fg=dark_txt,
                insertbackground=dark_txt,
                relief="flat",
                padx=10,
                pady=10,
                highlightthickness=1,
                highlightbackground=dark_bg,
            )
            txt.pack(fill="both", expand=True)

            for line in safe_lines:
                txt.insert(tk.END, f"{bullet} {line}\n\n")

            def _on_mousewheel_windows(event: tk.Event) -> str:
                txt.yview_scroll(int(-1 * (event.delta // 120)), "units")
                return "break"

            def _on_mousewheel_linux(event: tk.Event) -> str:
                direction = -1 if event.num == 4 else 1
                txt.yview_scroll(direction, "units")
                return "break"

            txt.bind("<MouseWheel>", _on_mousewheel_windows)
            txt.bind("<Button-4>", _on_mousewheel_linux)
            txt.bind("<Button-5>", _on_mousewheel_linux)

            btns = tk.Frame(alert, bg=dark_bg)
            btns.grid(row=2, column=0, sticky="ew", padx=22, pady=(8, 18))
            btns.grid_columnconfigure(0, weight=0)
            btns.grid_columnconfigure(1, weight=1)
            btns.grid_columnconfigure(2, weight=0)

            toast = tk.Label(
                btns,
                text="",
                font=("Segoe UI", 9),
                fg="#00c853",
                bg=dark_bg,
            )
            toast.grid(row=0, column=1, sticky="w", padx=10)
            toast.grid_remove()

            def _toast(msg: str, ms: int = 1200) -> None:
                toast.config(text=msg)
                toast.grid()
                alert.after(ms, toast.grid_remove)

            def copy_all() -> None:
                alert.clipboard_clear()
                alert.clipboard_append("\n".join(safe_lines))
                _toast("✅ Copied")

            copy_btn = tk.Button(
                btns,
                text="📋 Copy All",
                font=("Segoe UI", 10, "bold"),
                bg=accent_color,
                fg="white",
                relief="flat",
                padx=12,
                pady=4,
                command=copy_all,
            )
            copy_btn.grid(row=0, column=0, sticky="w")

            close_btn = tk.Button(
                btns,
                text="✖ Close",
                font=("Segoe UI", 10, "bold"),
                bg="#b30000",
                fg="white",
                relief="flat",
                padx=12,
                pady=4,
                command=alert.destroy,
            )
            close_btn.grid(row=0, column=2, sticky="e")

            font = tkfont.Font(font=("Segoe UI", 10))
            longest_px = 0
            for line in safe_lines:
                px = font.measure(f"{bullet} {line}")
                if px > longest_px:
                    longest_px = px

            desired_w = int(longest_px * 1.3) + (22 + 22) + 20
            line_h = max(18, font.metrics("linespace") + 2)
            n_items = max(1, len(safe_lines))
            desired_h = (18 + 12) + n_items * (line_h + 4) + (8 + 42 + 18)

            sw, sh = alert.winfo_screenwidth(), alert.winfo_screenheight()
            min_w, min_h = 520, 260
            max_w, max_h = int(sw * 0.80), int(sh * 0.80)

            win_w = max(min_w, min(max_w, desired_w))
            win_h = max(min_h, min(max_h, desired_h))

            alert.geometry(f"{win_w}x{win_h}+0+0")

            def _focus_after_map(_e=None) -> None:
                try:
                    txt.focus_set()
                except Exception:
                    pass

            alert.bind("<Visibility>", _focus_after_map)
            alert.after_idle(_focus_after_map)

            try:
                alert.attributes("-topmost", True)
                alert.after(200, lambda: alert.attributes("-topmost", False))
            except Exception:
                pass

            alert.wait_window()

        for file_path in csv_files:
            try:
                non_eu, eu, weight, box5, invoice, missing_coo = analyze_invoice(
                    file_path
                )
            except Exception as e:
                file_issues.append(
                    f"{os.path.basename(file_path)} → Unexpected error: {e}"
                )
                if logger:
                    logger.error(
                        "analyze_invoice crashed for %s: %s", file_path, e
                    )
                continue

            all_non_eu_items.extend(non_eu)
            all_eu_items.extend(eu)

            if box5 != "Unknown":
                all_po_numbers.append(box5)
            if invoice != "Unknown":
                all_invoice_numbers.append(invoice)

            total_weight += weight

            if missing_coo:
                errs = [
                    x
                    for x in missing_coo
                    if str(x).startswith("Error during analysis:")
                ]
                misses = [
                    x
                    for x in missing_coo
                    if not str(x).startswith("Error during analysis:")
                ]

                file_issues.extend(
                    [
                        f"{os.path.basename(file_path)} → {e}"
                        for e in errs
                    ]
                )

                if misses:
                    if multiple_files:
                        all_missing_coo.append(
                            f"📄 {os.path.basename(file_path)}"
                        )

                    for entry in misses:
                        core = entry.replace("⚠️ Missing COO → ", "")
                        parts = [p.strip() for p in core.split("\n")]

                        line_num = ""
                        product = ""
                        desc = ""
                        weight_txt = ""

                        for p in parts:
                            up = p.upper()
                            if up.startswith("LINE"):
                                line_num = p.split(" ", 1)[1].strip()
                            elif up.startswith("PRODUCT:"):
                                product = p.split(":", 1)[1].strip()
                            elif up.startswith("DESC:"):
                                desc = p.split(":", 1)[1].strip()
                            elif up.startswith("WEIGHT:"):
                                weight_txt = p.split(":", 1)[1].strip()

                        if "/" in weight_txt:
                            weight_txt = weight_txt.split("/")[-1].strip()
                        weight_txt = " ".join(weight_txt.split())

                        all_missing_coo.append(
                            f"Line {line_num} – {product}, {desc}, {weight_txt}"
                        )

        if logger:
            try:
                logger.info(
                    "Totals → NonEU=%d, EU=%d, POs=%d, Invoices=%d, Weight=%.3fKG, "
                    "MissingCOO=%d, Issues=%d",
                    len(all_non_eu_items),
                    len(all_eu_items),
                    len(set(all_po_numbers)),
                    len(set(all_invoice_numbers)),
                    round(total_weight, 3),
                    len(all_missing_coo),
                    len(file_issues),
                )
            except Exception:
                pass

        box5_list = (
            ", ".join(sorted(set(all_po_numbers))) if all_po_numbers else "Unknown"
        )
        invoice_list = (
            ", ".join(sorted(set(all_invoice_numbers)))
            if all_invoice_numbers
            else "Unknown"
        )

        message = (
            "Dear Team,\n\n"
            "Please proceed requesting the respective Certificate of Origin for the attached invoices.\n\n"
        )

        message += "Non-EU items:\n"
        for i, item in enumerate(all_non_eu_items, 1):
            message += f"{i}. {item}\n"

        message += f"\nReference: inv. {invoice_list}\n"
        message += f"Weight: {round(total_weight, 3)} KG\n"
        message += f"Box 5: {box5_list}\n\n"

        message += "EU items:\n"
        for i, item in enumerate(sorted(set(all_eu_items)), 1):
            message += f"{i}. {item}\n"

        message += (
            "\nBest regards,\n"
            "Customer Care Team\n"
            f"Email: {DEFAULT_SUPPORT_EMAIL}\n"
            f"Phone: {DEFAULT_SUPPORT_PHONE_URL}\n"
        )

        increment_invoice_analysis_counter()
        updated_count = get_invoice_analysis_count()
        elapsed_time = round(time.time() - start_time, 2)

        processing_label.config(
            text=(
                "⏱️ Processing Completed\n"
                f"Duration: {elapsed_time} seconds\n"
                f"Total Invoices Analyzed: {updated_count}"
            )
        )

        text_area.configure(state="normal")
        text_area.delete("1.0", tk.END)
        text_area.insert(tk.END, message)
        text_area.configure(state="normal")

        if all_missing_coo:
            if logger:
                logger.debug(
                    "Showing Missing COO modal; items=%d", len(all_missing_coo)
                )
            _show_list_modal(
                "Missing COO encountered — manual check required",
                all_missing_coo,
                bullet="🔸",
            )

        if file_issues:
            if logger:
                logger.debug(
                    "Showing File Issues modal; items=%d", len(file_issues)
                )
            _show_list_modal(
                "⚠️ File Issues detected — please review",
                file_issues,
                bullet="❗",
            )

        if enable_excel_export.get():
            user_name, should_open_excel = get_responsible_person()
            if not user_name:
                processing_label.config(
                    text="Tracking list generation: Cancelled by user"
                )
                return

            run_metadata_export(file_paths, user_name, should_open_excel)

    header_frame = tk.Frame(root)
    header_frame.pack(pady=20)

    button_frame = tk.Frame(root)
    button_frame.pack(pady=10)

    info_frame = tk.Frame(root)
    info_frame.pack(pady=10)

    footer_frame = tk.Frame(root)
    footer_frame.pack(pady=10)

    label = tk.Label(
        header_frame,
        text="📄 Invoice Analysis Console",
        font=("Segoe UI", 22, "bold"),
    )
    label.pack()

    subheader = tk.Label(
        header_frame,
        text="SmartDocs Insight v2.2 – Powered by Damian Komorowski",
        font=("Segoe UI", 8, "italic"),
        fg="#555555",
    )
    subheader.pack(pady=(5, 10))

    button = tk.Button(
        button_frame,
        text="📁 Choose Invoices to Analyze",
        command=select_files,
        font=("Segoe UI", 12, "bold"),
        padx=20,
        pady=10,
    )
    button.pack()

    excel_checkbox = tk.Checkbutton(
        button_frame,
        text="📋 Excel-Based Tracking Sheet Export",
        variable=enable_excel_export,
        font=("Segoe UI", 10, "bold"),
    )
    excel_checkbox.pack(pady=5)

    doc_button = tk.Button(
        button_frame,
        text="📊 Documentation Report",
        command=open_documentation,
        font=("Segoe UI", 10),
        relief="ridge",
    )
    doc_button.pack(pady=(10, 0))

    summary_label = tk.Label(
        info_frame,
        text="📌 Summary includes PO numbers, weights, and item details.",
        font=("Segoe UI", 12, "bold"),
    )
    summary_label.grid(row=0, column=0, sticky="w", padx=10)

    email_label = tk.Label(
        info_frame,
        text=f"Send to: {DEFAULT_COO_REQUEST_EMAIL}",
        font=("Segoe UI", 10),
    )
    email_label.grid(row=1, column=0, sticky="w", padx=10, pady=(5, 0))

    copy_button = tk.Button(
        info_frame,
        text="Copy",
        command=copy_email,
        font=("Segoe UI", 9),
        relief="ridge",
        padx=2,
    )
    copy_button.grid(row=1, column=1, sticky="w", padx=5)

    text_frame = tk.Frame(root)
    text_frame.pack(padx=20, pady=20, fill="both", expand=True)

    v_scrollbar = ttk.Scrollbar(text_frame, orient="vertical")
    v_scrollbar.pack(side="right", fill="y")

    text_area = tk.Text(
        text_frame,
        wrap="word",
        font=("Segoe UI", 10),
        yscrollcommand=v_scrollbar.set,
    )
    text_area.pack(side="left", fill="both", expand=True)
    v_scrollbar.config(command=text_area.yview)

    def copy_output() -> None:
        content = text_area.get("1.0", tk.END)
        root.clipboard_clear()
        root.clipboard_append(content)

    copy_output_button = tk.Button(
        root,
        text="📋 Copy Output",
        command=copy_output,
        font=("Segoe UI", 9),
        relief="ridge",
        padx=10,
    )
    copy_output_button.pack(pady=(0, 10))

    toggle_button = tk.Button(
        footer_frame,
        text="☀️ Light Mode" if dark_mode else "🌙 Dark Mode",
        command=toggle_dark_mode,
        font=("Segoe UI", 9),
        relief="ridge",
        padx=10,
    )
    toggle_button.pack(pady=5)

    analysis_count = get_invoice_analysis_count()
    processing_label = tk.Label(
        footer_frame,
        text=(
            "Status: Awaiting File Selection\n"
            f"Total Invoices Processed: {analysis_count}"
        ),
        font=("Segoe UI", 8),
    )
    processing_label.pack()

    apply_theme()
    root.mainloop()


# ---------------------------------------------------------------------------
# Entrypoint
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    show_gui()